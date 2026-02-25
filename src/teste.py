import pandas as pd
import dateutil.relativedelta 
import numpy as np
import openpyxl
import sys
import statsmodels.api as sm
from datetime import datetime
from statsmodels.graphics.tsaplots import plot_acf, plot_pacf
from statsmodels.tsa.statespace.sarimax import SARIMAX
import matplotlib.pyplot as plt

arquivo25 = r"C:\Users\izabe\Downloads\PEDIDOS X DIAS 2025.xlsx"
abas = ['AAE2', 'CANT2', 'PORT2', 'SERV2']

# Dicionário para armazenar os resultados de cada aba
resultados_abas = {}

# Processar cada aba separadamente
for aba in abas:
    print(f"\n{'='*60}")
    print(f"Processando aba: {aba}")
    print(f"{'='*60}\n")
    
    try:
        df_24 = pd.read_excel(arquivo25, sheet_name=aba)
    except Exception as e:
        print(f'erro ao ler a aba {aba}: {e}')
        continue
    
    colunas_fixas = ['DATA']
    
    # Obter índice dinâmico de B1 até a primeira coluna livre - 1
    wb = openpyxl.load_workbook(arquivo25)
    ws = wb[aba]
    
    # Encontrar primeira coluna vazia a partir de B (coluna 2)
    primeira_coluna_vazia = None
    for col_idx in range(2, ws.max_column + 2):
        if ws.cell(row=1, column=col_idx).value is None:
            primeira_coluna_vazia = col_idx
            break
    
    # Criar índice de B até primeira_coluna_vazia - 1
    if primeira_coluna_vazia:
        var_name = [ws.cell(row=1, column=col).value for col in range(2, primeira_coluna_vazia)]
    else:
        var_name = [ws.cell(row=1, column=col).value for col in range(2, ws.max_column + 1)]
    
    wb.close()
    
    # Converter para formato longo
    df_longo = df_24.melt(id_vars=colunas_fixas, value_vars=var_name, var_name='Escola', value_name='Valor')
    df_longo['Regiao'] = df_longo['Escola'].astype(str).str.extract(r'/([^/]+)/').str.strip()
    
    # Garantir que DATA é datetime
    df_longo['DATA'] = pd.to_datetime(df_longo['DATA'])
    
    # Agrupar por DATA e REGIAO, somando os valores
    df_agrupado = df_longo.groupby(['DATA', 'Regiao'])['Valor'].sum().reset_index()
    
    # Pivotar para ter regionais como colunas
    df_pivotado = df_agrupado.pivot(index='DATA', columns='Regiao', values='Valor').fillna(0)
    df_pivotado = df_pivotado.reset_index()
    
    # Criar range de todas as datas de 2025
    data_inicio = datetime(2025, 1, 1)
    data_fim = datetime(2025, 12, 31)
    todas_datas = pd.date_range(start=data_inicio, end=data_fim, freq='D')
    
    # Criar DataFrame com todas as datas
    df_final = pd.DataFrame({'DATA': todas_datas})
    
    # Fazer merge com os dados agrupados
    df_final = df_final.merge(df_pivotado, on='DATA', how='left').fillna(0)
    
    # Formatar a coluna DATA como DD/MM/YYYY
    df_final['DATA'] = df_final['DATA'].dt.strftime('%d/%m/%Y')
    
    df_final.to_excel(f'resultado_por_regiao2025_{aba}.xlsx', index=False)
    print(f"Arquivo 2025 para aba {aba} gerado com sucesso!")
    print(f"Total de linhas: {len(df_final)}")